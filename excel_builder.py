"""Builds the formatted 3-sheet workbook (J.V. / Reconciliation / Workings)
matching the user's Book5 presentation: navy headers, Cambria font, thin
borders on every cell, accounting number format, live formulas, and yellow
highlighting for categories flagged uncertain.
"""
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY = "FF002060"
YELLOW = "FFF2CC"
GREEN = "C6EFCE"
ACCT = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)'
_thin = Side(style="thin")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _cambria(size=11, bold=False, white=False):
    return Font(name="Cambria", size=size, bold=bold,
                color=("FFFFFFFF" if white else "FF000000"))


def _calibri(bold=False):
    return Font(name="Calibri", size=11, bold=bold)


def build_workbook(transactions, purchases, return_bytes=True):
    """transactions: list of dicts with keys date, serno, details,
       reward_points, intl_amount, amount, category, uncertain.
       purchases: the PDF 'Purchases / Charges' figure (for reconciliation).
       Returns BytesIO (return_bytes) or the openpyxl Workbook."""
    navy = PatternFill("solid", fgColor=NAVY)
    yfill = PatternFill("solid", fgColor=YELLOW)
    gfill = PatternFill("solid", fgColor=GREEN)

    wb = Workbook()

    # ---------- Workings ----------
    wk = wb.active
    wk.title = "Workings"
    heads = ["Date", "SerNo.", "Particulars", " Reward \nPoints",
             "Intl.# \namount", "Amount", "As per books"]
    for i, h in enumerate(heads):
        c = wk.cell(3, 2 + i, h)
        c.font = _cambria(bold=True, white=True)
        c.fill = navy
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="center")
    for ri, t in enumerate(transactions):
        r = 4 + ri
        dt = t["date"]
        a = wk.cell(r, 2, dt if isinstance(dt, date) else dt)
        a.number_format = "d-mmm-yy"
        wk.cell(r, 3, int(t["serno"]) if str(t["serno"]).isdigit() else t["serno"])
        wk.cell(r, 4, t["details"])
        wk.cell(r, 5, t["reward_points"])
        wk.cell(r, 6, t["intl_amount"])
        g = wk.cell(r, 7, t["amount"])
        g.number_format = ACCT
        h = wk.cell(r, 8, t["category"])
        if t.get("uncertain"):
            h.fill = yfill  # flag for manual review
        for c in range(2, 9):
            wk.cell(r, c).font = _cambria()
            wk.cell(r, c).border = BORDER
    last = 3 + len(transactions)
    tr = last + 1
    wk.cell(tr, 4, "TOTAL").font = _cambria(bold=True)
    tg = wk.cell(tr, 7, f"=SUM(G4:G{last})" if transactions else 0)
    tg.number_format = ACCT
    tg.font = _cambria(bold=True)
    for c in range(2, 9):
        wk.cell(tr, c).border = BORDER
    for col, w in {"A": 3, "B": 17.9, "C": 12.9, "D": 43.9,
                   "E": 16.5, "F": 13.4, "G": 14, "H": 22}.items():
        wk.column_dimensions[col].width = w

    # ---------- J.V. ----------
    jv = wb.create_sheet("J.V.")
    for i, h in enumerate(["G.L.", "Frequency", "Amount"]):
        c = jv.cell(3, 1 + i, h)
        c.font = _cambria(13, bold=True, white=True)
        c.fill = navy
        c.border = BORDER
        c.alignment = Alignment(vertical="center")
    cats = sorted({t["category"] for t in transactions})
    for i, cat in enumerate(cats):
        r = 4 + i
        jv.cell(r, 1, cat).font = _cambria()
        jv.cell(r, 2,
                f'=COUNTIF(Workings!$H$4:$H${last},A{r})').font = _cambria()
        cc = jv.cell(r, 3,
                     f'=SUMIF(Workings!$H$4:$H${last},A{r},'
                     f'Workings!$G$4:$G${last})')
        cc.number_format = ACCT
        cc.font = _cambria()
        for c in range(1, 4):
            jv.cell(r, c).border = BORDER
    gr = 4 + len(cats)
    jv.cell(gr, 1, "Grand Total").font = _cambria(bold=True)
    jv.cell(gr, 2, f"=SUM(B4:B{gr-1})" if cats else 0).font = _cambria(bold=True)
    gc = jv.cell(gr, 3, f"=SUM(C4:C{gr-1})" if cats else 0)
    gc.number_format = ACCT
    gc.font = _cambria(bold=True)
    for c in range(1, 4):
        jv.cell(gr, c).border = BORDER
    for col, w in {"A": 24, "B": 21.2, "C": 14}.items():
        jv.column_dimensions[col].width = w

    # ---------- Reconciliation ----------
    rc = wb.create_sheet("Reconciliation")
    rng = f"Workings!G4:G{last}"

    def block(r, label, formula, fill=None, label_font=None):
        l = rc.cell(r, 2, label)
        l.font = label_font or _cambria()
        l.border = BORDER
        c = rc.cell(r, 3, formula)
        c.number_format = ACCT
        c.font = _calibri(bold=True)
        c.border = BORDER
        if fill:
            c.fill = fill

    h1 = rc.cell(2, 2, "RECONCILIATION CHECK")
    h1.font = _cambria(bold=True, white=True); h1.fill = navy; h1.border = BORDER
    h1c = rc.cell(2, 3, "Amount")
    h1c.font = _cambria(bold=True, white=True); h1c.fill = navy; h1c.border = BORDER
    block(3, "Gross debits (purchases)", f'=SUMIF({rng},">0")', yfill)
    block(4, "Gross credits (CR entries)", f'=SUMIF({rng},"<0")', yfill)
    block(5, "Net (debits + credits)", '=C3+C4', yfill)
    ph = rc.cell(7, 2, "PDF Statement Summary (entered for check)")
    ph.font = _cambria(bold=True, white=True); ph.fill = navy; ph.border = BORDER
    rc.cell(7, 3).border = BORDER
    block(8, "Purchases / Charges (per PDF)",
          round(purchases, 2) if purchases is not None else 0,
          label_font=_cambria(bold=True))
    block(9, "Difference: my debits vs PDF", '=C3-C8')
    l = rc.cell(10, 2, "Status"); l.font = _cambria(bold=True); l.border = BORDER
    st = rc.cell(10, 3, '=IF(ROUND(C9,2)=0,"MATCH","CHECK")')
    st.font = _calibri(bold=True); st.fill = gfill; st.border = BORDER
    for col, w in {"A": 3, "B": 43.1, "C": 16}.items():
        rc.column_dimensions[col].width = w

    # Order: J.V., Reconciliation, Workings
    wb._sheets.sort(key=lambda s: {"J.V.": 0, "Reconciliation": 1,
                                   "Workings": 2}[s.title])

    if return_bytes:
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    return wb


def build_combined_workbook(statements, return_bytes=True):
    """Combine several statements into one workbook.

    statements: list of dicts with keys:
        label        - short statement label, e.g. "Neetika may"
        transactions - list of transaction dicts (as in build_workbook)
        purchases    - the PDF 'Purchases / Charges' figure for that statement

    Layout:
      Workings       - all rows, with a leading 'Statement' column
      J.V.           - combined category summary across everything
      Reconciliation - one row per statement (gross debits vs PDF) + total
    """
    navy = PatternFill("solid", fgColor=NAVY)
    yfill = PatternFill("solid", fgColor=YELLOW)
    gfill = PatternFill("solid", fgColor=GREEN)

    wb = Workbook()

    # ---------- Workings (Statement col added at B) ----------
    wk = wb.active
    wk.title = "Workings"
    heads = ["Statement", "Date", "SerNo.", "Particulars", " Reward \nPoints",
             "Intl.# \namount", "Amount", "As per books"]
    for i, h in enumerate(heads):
        c = wk.cell(3, 2 + i, h)
        c.font = _cambria(bold=True, white=True)
        c.fill = navy
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="center")
    r = 4
    for stmt in statements:
        for t in stmt["transactions"]:
            wk.cell(r, 2, stmt["label"])
            a = wk.cell(r, 3, t["date"])
            a.number_format = "d-mmm-yy"
            wk.cell(r, 4, int(t["serno"]) if str(t["serno"]).isdigit() else t["serno"])
            wk.cell(r, 5, t["details"])
            wk.cell(r, 6, t["reward_points"])
            wk.cell(r, 7, t["intl_amount"])
            g = wk.cell(r, 8, t["amount"])
            g.number_format = ACCT
            h = wk.cell(r, 9, t["category"])
            if t.get("uncertain"):
                h.fill = yfill
            for c in range(2, 10):
                wk.cell(r, c).font = _cambria()
                wk.cell(r, c).border = BORDER
            r += 1
    last = r - 1
    tr = r
    wk.cell(tr, 5, "TOTAL").font = _cambria(bold=True)
    tg = wk.cell(tr, 8, f"=SUM(H4:H{last})" if last >= 4 else 0)
    tg.number_format = ACCT
    tg.font = _cambria(bold=True)
    for c in range(2, 10):
        wk.cell(tr, c).border = BORDER
    for col, w in {"A": 3, "B": 16, "C": 12, "D": 12.9, "E": 43.9,
                   "F": 12, "G": 13.4, "H": 14, "I": 22}.items():
        wk.column_dimensions[col].width = w

    cat_col, amt_col = "I", "H"
    rng_cat = f"Workings!${cat_col}$4:${cat_col}${last}"
    rng_amt = f"Workings!${amt_col}$4:${amt_col}${last}"
    stmt_col = "B"
    rng_stmt = f"Workings!${stmt_col}$4:${stmt_col}${last}"

    # ---------- J.V. (combined) ----------
    jv = wb.create_sheet("J.V.")
    for i, h in enumerate(["G.L.", "Frequency", "Amount"]):
        c = jv.cell(3, 1 + i, h)
        c.font = _cambria(13, bold=True, white=True)
        c.fill = navy
        c.border = BORDER
        c.alignment = Alignment(vertical="center")
    cats = sorted({t["category"] for s in statements for t in s["transactions"]})
    for i, cat in enumerate(cats):
        rr = 4 + i
        jv.cell(rr, 1, cat).font = _cambria()
        jv.cell(rr, 2, f'=COUNTIF({rng_cat},A{rr})').font = _cambria()
        cc = jv.cell(rr, 3, f'=SUMIF({rng_cat},A{rr},{rng_amt})')
        cc.number_format = ACCT
        cc.font = _cambria()
        for c in range(1, 4):
            jv.cell(rr, c).border = BORDER
    gr = 4 + len(cats)
    jv.cell(gr, 1, "Grand Total").font = _cambria(bold=True)
    jv.cell(gr, 2, f"=SUM(B4:B{gr-1})" if cats else 0).font = _cambria(bold=True)
    gc = jv.cell(gr, 3, f"=SUM(C4:C{gr-1})" if cats else 0)
    gc.number_format = ACCT
    gc.font = _cambria(bold=True)
    for c in range(1, 4):
        jv.cell(gr, c).border = BORDER
    for col, w in {"A": 24, "B": 21.2, "C": 14}.items():
        jv.column_dimensions[col].width = w

    # ---------- Reconciliation (per statement) ----------
    # Columns: B Statement | C Gross debits | D Net | E PDF Purchases |
    #          F Difference | G Status
    rc = wb.create_sheet("Reconciliation")
    rec_heads = ["Statement", "Gross debits", "Net (after CR)",
                 "PDF Purchases / Charges", "Difference", "Status"]
    for i, h in enumerate(rec_heads):
        c = rc.cell(2, 2 + i, h)
        c.font = _cambria(bold=True, white=True)
        c.fill = navy
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="center")
    rr = 3
    for stmt in statements:
        rc.cell(rr, 2, stmt["label"]).font = _cambria()
        gd = rc.cell(rr, 3,
                     f'=SUMIFS({rng_amt},{rng_stmt},B{rr},{rng_amt},">0")')
        nt = rc.cell(rr, 4, f'=SUMIFS({rng_amt},{rng_stmt},B{rr})')
        pv = rc.cell(rr, 5, round(stmt["purchases"], 2)
                     if stmt["purchases"] is not None else 0)
        df = rc.cell(rr, 6, f'=C{rr}-E{rr}')
        for col in (3, 4, 5, 6):
            rc.cell(rr, col).number_format = ACCT
            rc.cell(rr, col).font = _calibri()
        stt = rc.cell(rr, 7, f'=IF(ROUND(F{rr},2)=0,"MATCH","CHECK")')
        stt.font = _calibri(bold=True)
        stt.fill = gfill
        for c in range(2, 8):
            rc.cell(rr, c).border = BORDER
        rr += 1
    # Total row
    rc.cell(rr, 2, "TOTAL").font = _cambria(bold=True)
    for col, letter in ((3, "C"), (4, "D"), (5, "E"), (6, "F")):
        tc = rc.cell(rr, col, f"=SUM({letter}3:{letter}{rr-1})")
        tc.number_format = ACCT
        tc.font = _calibri(bold=True)
    allst = rc.cell(rr, 7, f'=IF(ROUND(F{rr},2)=0,"MATCH","CHECK")')
    allst.font = _calibri(bold=True)
    allst.fill = gfill
    for c in range(2, 8):
        rc.cell(rr, c).border = BORDER
    for col, w in {"A": 3, "B": 18, "C": 15, "D": 15, "E": 18,
                   "F": 13, "G": 11}.items():
        rc.column_dimensions[col].width = w

    wb._sheets.sort(key=lambda s: {"J.V.": 0, "Reconciliation": 1,
                                   "Workings": 2}[s.title])

    if return_bytes:
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    return wb
